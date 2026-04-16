import io
import re
import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates

from app.clients.itop import _call

router = APIRouter(prefix="/import/agrofert", tags=["agrofert-import"])
templates = Jinja2Templates(directory="app/templates")

# XLSX Typ požadavku → iTOP request_type
REQUEST_TYPE_MAP = {
    "chyba / paušál": "incident",
    "nový požadavek": "service_request",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _excel_serial_to_dt(serial) -> Optional[datetime.datetime]:
    """Convert Excel date serial number to datetime."""
    if not serial:
        return None
    try:
        return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(serial))
    except (TypeError, ValueError):
        return None


def _parse_date(text) -> Optional[datetime.datetime]:
    """Parse 'DD.MM.YYYY HH:MM' string to datetime."""
    if not text:
        return None
    if isinstance(text, datetime.datetime):
        return text
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(str(text).strip(), fmt)
        except ValueError:
            pass
    return None


def _parse_person(text: str) -> Optional[dict]:
    """Parse 'Příjmení Jméno (email@domain.cz)' → {last_name, first_name, email}."""
    if not text:
        return None
    m = re.match(r"^(.+?)\s*\(([^)]+@[^)]+)\)\s*$", text.strip())
    if not m:
        return None
    full_name = m.group(1).strip()
    email = m.group(2).strip().lower()
    parts = full_name.split()
    last_name = parts[0] if parts else full_name
    first_name = " ".join(parts[1:]) if len(parts) > 1 else ""
    return {"last_name": last_name, "first_name": first_name, "email": email}


def _get_org_id(person_email: str, agrofert_org_id: str, alintrust_org_id: str) -> str:
    """Determine org by email domain: @alintrust.cz → Alintrust, else → Agrofert."""
    if person_email.endswith("@alintrust.cz"):
        return alintrust_org_id
    return agrofert_org_id


def _find_or_create_person(person: dict, org_id: str) -> str:
    """Find Person by email; create if not found. Returns iTOP object id."""
    result = _call({
        "operation": "core/get",
        "class": "Person",
        "key": f"SELECT Person WHERE email = '{person['email']}'",
        "output_fields": "name",
        "limit": 1,
    })
    objs = result.get("objects") or {}
    if objs:
        return next(iter(objs.keys()))

    create_result = _call({
        "operation": "core/create",
        "class": "Person",
        "fields": {
            "name": person["last_name"],
            "first_name": person["first_name"],
            "email": person["email"],
            "org_id": f"SELECT Organization WHERE id = {org_id}",
        },
        "output_fields": "name",
    })
    objs = create_result.get("objects") or {}
    if not objs:
        raise RuntimeError(f"Nelze vytvořit Person: {person['email']}")
    return next(iter(objs.keys()))


def _find_contract_id(sluzba: str) -> Optional[str]:
    """Find CustomerContract by exact name."""
    result = _call({
        "operation": "core/get",
        "class": "CustomerContract",
        "key": f"SELECT CustomerContract WHERE name = '{sluzba}'",
        "output_fields": "name",
        "limit": 1,
    })
    objs = result.get("objects") or {}
    return next(iter(objs.keys())) if objs else None


def _ticket_exists(external_id: str) -> Optional[str]:
    """Return iTOP ticket id if UserRequest with given external_id exists, else None."""
    result = _call({
        "operation": "core/get",
        "class": "UserRequest",
        "key": f"SELECT UserRequest WHERE external_id = '{external_id}'",
        "output_fields": "ref",
        "limit": 1,
    })
    objs = result.get("objects") or {}
    return next(iter(objs.keys())) if objs else None


def _create_worklog(
    ticket_id: str,
    agent_id: str,
    start_dt: datetime.datetime,
    duration_h: float,
    description: str,
    is_billable: bool,
    work_location: str = "remote",
    kilometers: float = 0.0,
) -> str:
    duration_sec = int(duration_h * 3600)
    end_dt = start_dt + datetime.timedelta(seconds=duration_sec)
    fields = {
        "ticket_id": f"SELECT UserRequest WHERE id = {ticket_id}",
        "agent_id": f"SELECT Person WHERE id = {agent_id}",
        "start_date": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "end_date": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "duration": str(duration_sec),
        "input_mode": "duration",
        "description": description or "-",
        "is_billable": "yes" if is_billable else "no",
        "work_location": work_location,
    }
    if kilometers:
        fields["kilometers"] = str(kilometers)
    result = _call({
        "operation": "core/create",
        "class": "WorkLog",
        "fields": fields,
        "output_fields": "start_date",
    })
    objs = result.get("objects") or {}
    return next(iter(objs.keys())) if objs else ""


# ── Core import logic ─────────────────────────────────────────────────────────

def import_rows(
    rows: list[dict],
    agrofert_org_id: str,
    alintrust_org_id: str,
    dry_run: bool,
) -> list[dict]:
    results = []
    contract_cache: dict[str, Optional[str]] = {}
    person_cache: dict[str, str] = {}  # email → iTOP id

    for row in rows:
        external_id = str(row.get("ID") or "").strip()
        title = str(row.get("Předmět") or "").strip()
        sluzba = str(row.get("Služba") or "").strip()

        result_row = {
            "external_id": external_id or "?",
            "title": title,
            "sluzba": sluzba,
            "time_h": row.get("Odpracovaný čas (Př.: 1,5 = 1 hod 30 min; 0,25 = 15 min)"),
        }

        if not external_id:
            result_row.update(status="skip", note="prázdné ID")
            results.append(result_row)
            continue

        if dry_run:
            result_row.update(status="dry-run", note=(
                f"Společnost: {row.get('Společnost')} | "
                f"Autor: {row.get('Autor')} | "
                f"Odpracoval: {row.get('Odpracoval')}"
            ))
            results.append(result_row)
            continue

        try:
            # Deduplication check
            existing = _ticket_exists(external_id)
            if existing:
                result_row.update(status="skip", note=f"již existuje (iTOP id {existing})")
                results.append(result_row)
                continue

            # Resolve contract
            if sluzba not in contract_cache:
                contract_cache[sluzba] = _find_contract_id(sluzba)
            contract_id = contract_cache[sluzba]
            if not contract_id:
                result_row.update(status="error", note=f"kontrakt nenalezen: '{sluzba}'")
                results.append(result_row)
                continue

            # Resolve caller (zákazník)
            caller_id = None
            caller_raw = str(row.get("Autor") or "").strip()
            if caller_raw:
                p = _parse_person(caller_raw)
                if p:
                    if p["email"] not in person_cache:
                        org = _get_org_id(p["email"], agrofert_org_id, alintrust_org_id)
                        person_cache[p["email"]] = _find_or_create_person(p, org)
                    caller_id = person_cache[p["email"]]

            # Resolve agent (technik Alintrust)
            agent_id = None
            agent_raw = str(row.get("Odpracoval") or "").strip()
            if agent_raw:
                p = _parse_person(agent_raw)
                if p:
                    if p["email"] not in person_cache:
                        org = _get_org_id(p["email"], agrofert_org_id, alintrust_org_id)
                        person_cache[p["email"]] = _find_or_create_person(p, org)
                    agent_id = person_cache[p["email"]]

            # Dates
            close_dt = _parse_date(row.get("Datum vyřešení / zrušení / pozastavení"))
            worklog_dt = _excel_serial_to_dt(
                row.get("Datum vytvoření výkazu")
            ) or close_dt

            # Description: odkaz na AGF HD + dceřiná společnost
            odkaz = str(row.get("Odkaz na požadavek") or "").strip()
            subsidiary = str(row.get("Společnost") or "").strip()
            description_parts = [f"Dceřiná společnost: {subsidiary}"]
            if odkaz:
                description_parts.append(f'AGF HD: <a href="{odkaz}">{external_id}</a>')
            navazano = str(row.get("Navázáno na souvisejícím požadavek") or "").strip()
            if navazano:
                description_parts.append(f"Navázáno na: {navazano}")

            description = "<p>" + "<br/>".join(description_parts) + "</p>"

            # request_type mapping
            typ_raw = str(row.get("Typ požadavku") or "").strip().lower()
            request_type = REQUEST_TYPE_MAP.get(typ_raw, "service_request")

            # Placená služba → billable
            is_billable = str(row.get("Placená služba") or "").strip().lower() == "ano"

            # Build UserRequest fields
            fields: dict = {
                "title": title,
                "description": description,
                "org_id": f"SELECT Organization WHERE id = {agrofert_org_id}",
                "contract_id": f"SELECT CustomerContract WHERE id = {contract_id}",
                "status": "closed",
                "origin": "phone",
                "request_type": request_type,
                "external_id": external_id,
                "subsidiary": subsidiary,
            }
            if odkaz:
                fields["agf_ticket_url"] = odkaz
            jira_id = str(row.get("ID JIRA") or "").strip()
            if jira_id:
                fields["agf_jira_id"] = jira_id
            if caller_id:
                fields["caller_id"] = f"SELECT Person WHERE id = {caller_id}"
            if agent_id:
                fields["agent_id"] = f"SELECT Person WHERE id = {agent_id}"
            if close_dt:
                ts = close_dt.strftime("%Y-%m-%d %H:%M:%S")
                fields["close_date"] = ts
                fields["start_date"] = ts
                fields["resolution_date"] = ts

            create_result = _call({
                "operation": "core/create",
                "class": "UserRequest",
                "fields": fields,
                "output_fields": "ref",
            })
            ticket_objs = create_result.get("objects") or {}
            if not ticket_objs:
                raise RuntimeError("core/create nevrátil objekt")
            ticket_key = next(iter(ticket_objs.keys()))
            ticket_ref = next(iter(ticket_objs.values()))["fields"].get("ref", ticket_key)

            # WorkLog
            duration_h = float(
                row.get("Odpracovaný čas (Př.: 1,5 = 1 hod 30 min; 0,25 = 15 min)") or 0
            )
            comment = str(row.get("Komentář") or "").strip()
            km_raw = row.get("Kilometry")
            km_val = 0.0
            try:
                km_val = float(km_raw) if km_raw else 0.0
            except (TypeError, ValueError):
                pass
            if duration_h > 0 and agent_id and worklog_dt:
                _create_worklog(
                    ticket_id=ticket_key,
                    agent_id=agent_id,
                    start_dt=worklog_dt,
                    duration_h=duration_h,
                    description=comment or title,
                    is_billable=is_billable,
                    work_location="onsite" if km_val > 0 else "remote",
                    kilometers=km_val,
                )

            result_row.update(status="ok", note=f"vytvořen {ticket_ref}")

        except Exception as exc:
            result_row.update(status="error", note=str(exc))

        results.append(result_row)

    return results


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def import_form(request: Request):
    # Load available organizations for the form dropdowns
    try:
        org_result = _call({
            "operation": "core/get",
            "class": "Organization",
            "key": "SELECT Organization",
            "output_fields": "name",
            "limit": 0,
        })
        orgs = [
            {"id": k, "name": v["fields"]["name"]}
            for k, v in (org_result.get("objects") or {}).items()
        ]
        orgs.sort(key=lambda o: o["name"])
    except Exception:
        orgs = []

    return templates.TemplateResponse("agrofert_import.html", {
        "request": request,
        "orgs": orgs,
        "results": None,
        "stats": None,
    })


@router.post("/", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    file: UploadFile = File(...),
    agrofert_org_id: str = Form(...),
    alintrust_org_id: str = Form(...),
    dry_run: bool = Form(False),
):
    # Load orgs for re-render
    try:
        org_result = _call({
            "operation": "core/get",
            "class": "Organization",
            "key": "SELECT Organization",
            "output_fields": "name",
            "limit": 0,
        })
        orgs = [
            {"id": k, "name": v["fields"]["name"]}
            for k, v in (org_result.get("objects") or {}).items()
        ]
        orgs.sort(key=lambda o: o["name"])
    except Exception:
        orgs = []

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if any(row)
    ]

    results = import_rows(rows, agrofert_org_id, alintrust_org_id, dry_run)

    stats = {
        "ok":    sum(1 for r in results if r["status"] == "ok"),
        "skip":  sum(1 for r in results if r["status"] == "skip"),
        "error": sum(1 for r in results if r["status"] == "error"),
        "dry":   sum(1 for r in results if r["status"] == "dry-run"),
        "total": len(results),
    }

    return templates.TemplateResponse("agrofert_import.html", {
        "request": request,
        "orgs": orgs,
        "results": results,
        "stats": stats,
        "dry_run": dry_run,
        "filename": file.filename,
    })
