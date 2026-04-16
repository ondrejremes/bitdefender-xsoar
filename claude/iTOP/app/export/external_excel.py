"""
External customer report Excel builder.
Sheets: Souhrn, Servisní požadavky, Pravidelné aktivity, Výkaz práce, Čerpání paušálu
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Colors ──────────────────────────────────────────────────────────────────
C_DARK   = "1A2235"
C_PURPLE = "6B21A8"
C_LIGHT  = "EDE9FE"
C_GREEN  = "D1FAE5"
C_RED    = "FEE2E2"
C_GREY   = "F3F4F6"
C_WHITE  = "FFFFFF"

# ── Styles helpers ───────────────────────────────────────────────────────────
def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Calibri")
def _border(): return Border(bottom=Side(style="thin", color="D1D5DB"))
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _header_row(ws, cols, row=1, fill=C_DARK, font_color=C_WHITE, height=22):
    ws.row_dimensions[row].height = height
    for i, (col, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill = _fill(fill)
        c.font = _font(bold=True, color=font_color, size=9)
        c.alignment = _center()
        ws.column_dimensions[get_column_letter(i)].width = width

def _data_row(ws, values, row, shade=False):
    fill = _fill(C_GREY) if shade else _fill(C_WHITE)
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.fill = fill
        c.font = _font(size=9)
        c.alignment = _left()
        c.border = _border()

def _title_block(ws, title, subtitle, logo_path=None):
    """Add logo + title block to top of sheet."""
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 10

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.height = 50
            img.width = 180
            ws.add_image(img, "A1")
        except Exception:
            pass

    t = ws.cell(row=1, column=4, value=title)
    t.font = Font(bold=True, size=14, color=C_PURPLE, name="Calibri")
    t.alignment = _left()

    s = ws.cell(row=2, column=4, value=subtitle)
    s.font = _font(color="6B7280", size=10)
    s.alignment = _left()


PRIORITY_CS = {"1": "Kritická", "2": "Vysoká", "3": "Střední", "4": "Nízká"}
PRIORITY_EN = {"1": "Critical", "2": "High", "3": "Medium", "4": "Low"}
STATUS_CS = {"new": "Nový", "assigned": "Přiřazený", "closed": "Uzavřený"}
STATUS_EN = {"new": "New", "assigned": "Assigned", "closed": "Closed"}


def _t(key, lang):
    """Simple translation helper."""
    translations = {
        "summary":          {"cs": "Souhrn",                    "en": "Summary"},
        "service_requests": {"cs": "Servisní požadavky",        "en": "Service Requests"},
        "gen_activities":   {"cs": "Pravidelné aktivity",       "en": "Recurring Activities"},
        "work_log":         {"cs": "Výkaz práce",               "en": "Work Log"},
        "budget":           {"cs": "Čerpání paušálu",           "en": "Budget Consumption"},
        "customer":         {"cs": "Zákazník",                  "en": "Customer"},
        "contract":         {"cs": "Kontrakt",                  "en": "Contract"},
        "period":           {"cs": "Období",                    "en": "Period"},
        "total_h":          {"cs": "Celkem hodin",              "en": "Total Hours"},
        "billable_h":       {"cs": "Fakturovatelné hodiny",     "en": "Billable Hours"},
        "non_billable_h":   {"cs": "Nefakturovatelné hodiny",   "en": "Non-billable Hours"},
        "tickets_count":    {"cs": "Počet ticketů",             "en": "Ticket Count"},
        "ref":              {"cs": "Evidenční číslo",           "en": "Reference"},
        "date_open":        {"cs": "Datum otevření",            "en": "Date Opened"},
        "date_close":       {"cs": "Datum uzavření",            "en": "Date Closed"},
        "description":      {"cs": "Popis požadavku",           "en": "Description"},
        "priority":         {"cs": "Priorita",                  "en": "Priority"},
        "caller":           {"cs": "Zadavatel",                 "en": "Requester"},
        "status":           {"cs": "Stav",                      "en": "Status"},
        "agent":            {"cs": "Technik",                   "en": "Technician"},
        "response_time":    {"cs": "Response time [hod]",       "en": "Response Time [hrs]"},
        "resolution_time":  {"cs": "Resolution time [hod]",     "en": "Resolution Time [hrs]"},
        "work_type":        {"cs": "Typ činnosti",              "en": "Work Type"},
        "work_desc":        {"cs": "Popis činnosti",            "en": "Activity Description"},
        "hours":            {"cs": "Hodiny",                    "en": "Hours"},
        "billable":         {"cs": "V rámci paušálu",          "en": "Within Flat Rate"},
        "date":             {"cs": "Datum",                     "en": "Date"},
        "month":            {"cs": "Měsíc",                     "en": "Month"},
        "budget_h":         {"cs": "Rozpočet hodin",            "en": "Budget Hours"},
        "consumed_h":       {"cs": "Čerpáno hodin",             "en": "Consumed Hours"},
        "remaining_h":      {"cs": "Zbývá hodin",               "en": "Remaining Hours"},
        "yes":              {"cs": "Ano",                       "en": "Yes"},
        "no":               {"cs": "Ne",                        "en": "No"},
        "sla_ok":           {"cs": "Splněno",                   "en": "Met"},
        "sla_breach":       {"cs": "Porušeno",                  "en": "Breached"},
        "ticket":           {"cs": "Ticket",                    "en": "Ticket"},
        "workorder":        {"cs": "Work Order",                "en": "Work Order"},
        "gen_task":         {"cs": "Pravidelná aktivita",       "en": "Recurring Activity"},
        "worklog_id":       {"cs": "ID výkazu",                 "en": "Worklog ID"},
    }
    return translations.get(key, {}).get(lang, key)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, contract, tickets, date_from, date_to, lang, logo_path):
    ws = wb.active
    ws.title = _t("summary", lang)

    portal_tickets = [t for t in tickets if not t["is_general_task"]]
    all_worklogs = [wl for t in tickets for wo in t["workorders"] for wl in wo["worklogs"]]
    total_h = round(sum(wl["duration_h"] for wl in all_worklogs), 2)
    billable_h = round(sum(wl["duration_h"] for wl in all_worklogs if wl["is_billable"]), 2)

    subtitle = f"{contract['org']}  |  {contract['name']}  |  {date_from} – {date_to}"
    _title_block(ws, _t("summary", lang).upper(), subtitle, logo_path)

    # Summary table
    row = 6
    header = ws.cell(row=row, column=1, value=_t("summary", lang))
    header.font = _font(bold=True, color=C_WHITE, size=10)
    header.fill = _fill(C_PURPLE)
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 20
    row += 1

    data = [
        (_t("customer", lang),       contract["org"]),
        (_t("contract", lang),       contract["name"]),
        (_t("period", lang),         f"{date_from} – {date_to}"),
        (_t("total_h", lang),        total_h),
        (_t("billable_h", lang),     billable_h),
        (_t("non_billable_h", lang), round(total_h - billable_h, 2)),
        (_t("tickets_count", lang),  len(portal_tickets)),
    ]
    for i, (label, value) in enumerate(data):
        shade = i % 2 == 0
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = _font(bold=True, size=9)
        vc.font = _font(size=9)
        lc.fill = vc.fill = _fill(C_GREY if shade else C_WHITE)
        lc.alignment = _left()
        vc.alignment = _left()
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    # Header cell alignment
    header.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _sheet_service_requests(wb, tickets, lang, logo_path):
    ws = wb.create_sheet(_t("service_requests", lang))
    portal = [t for t in tickets if not t["is_general_task"]]

    _title_block(ws, _t("service_requests", lang).upper(), "", logo_path)

    prio = PRIORITY_CS if lang == "cs" else PRIORITY_EN
    stat = STATUS_CS if lang == "cs" else STATUS_EN

    cols = [
        (_t("ref", lang), 14),
        (_t("date_open", lang), 18),
        (_t("date_close", lang), 18),
        (_t("description", lang), 40),
        (_t("priority", lang), 12),
        (_t("caller", lang), 18),
        (_t("status", lang), 14),
        (_t("agent", lang), 18),
        (_t("response_time", lang), 16),
        (_t("resolution_time", lang), 16),
        (_t("work_type", lang), 18),
        (_t("worklog_id", lang), 12),
        (_t("work_desc", lang), 45),
        (_t("hours", lang), 9),
        (_t("billable", lang), 14),
    ]
    _header_row(ws, cols, row=6)

    r = 7
    for t in sorted(portal, key=lambda x: x["start_date"]):
        worklogs = [(wo["name"], wl) for wo in t["workorders"] for wl in wo["worklogs"]]
        if not worklogs:
            worklogs = [("", None)]

        for idx, (wo_name, wl) in enumerate(worklogs):
            shade = (r % 2 == 0)
            values = [
                t["ref"] if idx == 0 else "",
                t["start_date"][:10] if idx == 0 else "",
                t["close_date"][:10] if t["close_date"] and idx == 0 else "",
                t["title"] if idx == 0 else "",
                prio.get(t["priority"], t["priority"]) if idx == 0 else "",
                t["caller"] if idx == 0 else "",
                stat.get(t["status"], t["status"]) if idx == 0 else "",
                wl["agent"] if wl else (t["agent"] if idx == 0 else ""),
                t["tto_h"] if t["tto_h"] is not None and idx == 0 else "",
                t["ttr_h"] if t["ttr_h"] is not None and idx == 0 else "",
                wo_name,
                wl["id"] if wl else "",
                wl["description"] if wl else "",
                wl["duration_h"] if wl else "",
                _t("yes", lang) if wl and wl["is_billable"] else (_t("no", lang) if wl else ""),
            ]
            _data_row(ws, values, r, shade)

            # SLA color for response/resolution
            if idx == 0:
                tto_cell = ws.cell(row=r, column=9)
                ttr_cell = ws.cell(row=r, column=10)
                if t["sla_tto_passed"]:
                    tto_cell.fill = _fill(C_RED)
                if t["sla_ttr_passed"]:
                    ttr_cell.fill = _fill(C_RED)


            r += 1


def _sheet_gen_activities(wb, tickets, lang, logo_path):
    ws = wb.create_sheet(_t("gen_activities", lang))
    gen = [t for t in tickets if t["is_general_task"]]

    _title_block(ws, _t("gen_activities", lang).upper(), "", logo_path)

    cols = [
        (_t("ref", lang), 14),
        (_t("date_open", lang), 18),
        (_t("description", lang), 40),
        (_t("workorder", lang), 22),
        (_t("worklog_id", lang), 12),
        (_t("work_desc", lang), 45),
        (_t("agent", lang), 18),
        (_t("hours", lang), 9),
        (_t("billable", lang), 14),
    ]
    _header_row(ws, cols, row=6)

    r = 7
    for t in sorted(gen, key=lambda x: x["start_date"]):
        worklogs = [(wo["name"], wl) for wo in t["workorders"] for wl in wo["worklogs"]]
        if not worklogs:
            worklogs = [("", None)]
        for idx, (wo_name, wl) in enumerate(worklogs):
            shade = r % 2 == 0
            values = [
                t["ref"] if idx == 0 else "",
                t["start_date"][:10] if idx == 0 else "",
                t["title"] if idx == 0 else "",
                wo_name,
                wl["id"] if wl else "",
                wl["description"] if wl else "",
                wl["agent"] if wl else "",
                wl["duration_h"] if wl else "",
                _t("yes", lang) if wl and wl["is_billable"] else (_t("no", lang) if wl else ""),
            ]
            _data_row(ws, values, r, shade)
            r += 1


def _sheet_work_log(wb, tickets, lang, logo_path):
    ws = wb.create_sheet(_t("work_log", lang))
    _title_block(ws, _t("work_log", lang).upper(), "", logo_path)

    cols = [
        (_t("date", lang), 14),
        (_t("ticket", lang), 14),
        (_t("work_type", lang), 22),
        (_t("worklog_id", lang), 12),
        (_t("work_desc", lang), 50),
        (_t("agent", lang), 18),
        (_t("hours", lang), 9),
        (_t("billable", lang), 14),
    ]
    _header_row(ws, cols, row=6)

    # Flatten all worklogs sorted by date
    entries = []
    for t in tickets:
        for wo in t["workorders"]:
            for wl in wo["worklogs"]:
                entries.append({
                    "id": wl["id"],
                    "date": wl["start_date"][:10],
                    "ref": t["ref"],
                    "wo_name": wo["name"],
                    "description": wl["description"],
                    "agent": wl["agent"],
                    "hours": wl["duration_h"],
                    "billable": wl["is_billable"],
                })
    entries.sort(key=lambda x: x["date"])

    r = 7
    for i, e in enumerate(entries):
        _data_row(ws, [
            e["date"], e["ref"], e["wo_name"], e["id"], e["description"],
            e["agent"], e["hours"],
            _t("yes", lang) if e["billable"] else _t("no", lang),
        ], r, shade=i % 2 == 0)
        r += 1

    # Total row
    total = round(sum(e["hours"] for e in entries), 2)
    tc = ws.cell(row=r, column=6, value="CELKEM" if lang == "cs" else "TOTAL")
    tc.font = _font(bold=True)
    vc = ws.cell(row=r, column=7, value=total)
    vc.font = _font(bold=True)


def _sheet_budget(wb, contract, tickets, date_from, date_to, lang, logo_path):
    if contract.get("budget_mode") == "none":
        return

    ws = wb.create_sheet(_t("budget", lang))
    _title_block(ws, _t("budget", lang).upper(), "", logo_path)

    cols = [
        (_t("month", lang), 18),
        (_t("budget_h", lang), 16),
        (_t("consumed_h", lang), 16),
        (_t("remaining_h", lang), 16),
    ]
    _header_row(ws, cols, row=6)

    r = 7
    for i, mb in enumerate(contract.get("monthly_budgets", [])):
        period = mb.get("period_start", "")[:7]
        if period < date_from[:7] or period > date_to[:7]:
            continue
        _data_row(ws, [
            period,
            float(mb.get("budgeted_hours") or 0),
            float(mb.get("consumed_hours") or 0),
            float(mb.get("remaining_hours") or 0),
        ], r, shade=i % 2 == 0)
        r += 1


# ── Main entry point ──────────────────────────────────────────────────────────

def build_external_excel(contract, tickets, date_from, date_to, lang="cs",
                         show_general=True, logo_path=None) -> bytes:
    if not show_general:
        display_tickets = tickets
    else:
        display_tickets = tickets

    wb = Workbook()
    _sheet_summary(wb, contract, display_tickets, date_from, date_to, lang, logo_path)
    _sheet_service_requests(wb, display_tickets, lang, logo_path)
    _sheet_gen_activities(wb, display_tickets, lang, logo_path)
    _sheet_work_log(wb, display_tickets, lang, logo_path)
    _sheet_budget(wb, contract, display_tickets, date_from, date_to, lang, logo_path)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
