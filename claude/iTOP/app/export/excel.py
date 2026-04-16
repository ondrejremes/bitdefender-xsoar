import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1A2235")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8EDF5")
SUBHEADER_FONT = Font(bold=True, size=10)
TOTAL_FONT = Font(bold=True)
BORDER_THIN = Border(
    bottom=Side(style="thin", color="DEE2E6"),
)


def _header(ws, cols: list[str], row: int = 1):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _sheet_summary(wb, date_from, date_to, total_h, total_billable_h, worklogs):
    ws = wb.active
    ws.title = "Souhrn"

    ws["A1"] = "Interní reporting"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Období: {date_from} – {date_to}"
    ws["A2"].font = Font(size=10, color="666666")
    ws.append([])

    _header(ws, ["Ukazatel", "Hodnota"], row=4)
    rows = [
        ("Celkem hodin", total_h),
        ("Fakturovatelné hodiny", round(total_billable_h, 2)),
        ("Nefakturovatelné hodiny", round(total_h - total_billable_h, 2)),
        ("Počet záznamů", len(worklogs)),
    ]
    for r in rows:
        ws.append(r)

    _autowidth(ws)


def _sheet_by_agent(wb, by_agent):
    ws = wb.create_sheet("Podle technika")
    _header(ws, ["Technik", "Celkem hodin", "Fakturovatelné", "Nefakturovatelné", "Počet záznamů"])
    for row in by_agent:
        ws.append([
            row["agent"],
            row["hours"],
            row["hours_billable"],
            round(row["hours"] - row["hours_billable"], 2),
            row["entries"],
        ])
    # Total row
    totals = ws.append([
        "CELKEM",
        sum(r["hours"] for r in by_agent),
        sum(r["hours_billable"] for r in by_agent),
        sum(round(r["hours"] - r["hours_billable"], 2) for r in by_agent),
        sum(r["entries"] for r in by_agent),
    ])
    last = ws.max_row
    for cell in ws[last]:
        cell.font = TOTAL_FONT
    _autowidth(ws)


def _sheet_by_contract(wb, by_contract):
    ws = wb.create_sheet("Podle kontraktu")
    _header(ws, ["Kontrakt", "Zákazník", "Celkem hodin", "Fakturovatelné", "Nefakturovatelné", "Počet záznamů"])
    for row in by_contract:
        ws.append([
            row["contract"],
            row["org"],
            row["hours"],
            row["hours_billable"],
            round(row["hours"] - row["hours_billable"], 2),
            row["entries"],
        ])
    last = ws.max_row + 1
    ws.append([
        "CELKEM", "",
        sum(r["hours"] for r in by_contract),
        sum(r["hours_billable"] for r in by_contract),
        sum(round(r["hours"] - r["hours_billable"], 2) for r in by_contract),
        sum(r["entries"] for r in by_contract),
    ])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    _autowidth(ws)


def _sheet_by_workorder(wb, by_workorder):
    ws = wb.create_sheet("Podle typu práce")
    _header(ws, ["Typ práce (WorkOrder)", "Celkem hodin", "Fakturovatelné", "Počet záznamů"])
    for row in by_workorder:
        ws.append([row["workorder"], row["hours"], row["hours_billable"], row["entries"]])
    ws.append([
        "CELKEM",
        sum(r["hours"] for r in by_workorder),
        sum(r["hours_billable"] for r in by_workorder),
        sum(r["entries"] for r in by_workorder),
    ])
    for cell in ws[ws.max_row]:
        cell.font = TOTAL_FONT
    _autowidth(ws)


def _sheet_detail(wb, worklogs):
    ws = wb.create_sheet("Detail záznamů")
    _header(ws, ["Datum", "Technik", "Zákazník", "Kontrakt", "WorkOrder", "Ticket", "Popis", "Hodiny", "Fakturovatelné"])
    for wl in worklogs:
        ws.append([
            wl["start_date"][:10],
            wl["agent"],
            wl["org"],
            wl["contract"],
            wl["workorder"],
            wl["ticket_ref"],
            wl["description"],
            wl["duration_h"],
            "Ano" if wl["is_billable"] else "Ne",
        ])
    _autowidth(ws)


def build_excel(date_from, date_to, total_h, total_billable_h,
                by_agent, by_contract, by_workorder, worklogs) -> bytes:
    wb = Workbook()
    _sheet_summary(wb, date_from, date_to, total_h, total_billable_h, worklogs)
    _sheet_by_agent(wb, by_agent)
    _sheet_by_contract(wb, by_contract)
    _sheet_by_workorder(wb, by_workorder)
    _sheet_detail(wb, worklogs)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
